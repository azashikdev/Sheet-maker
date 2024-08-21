import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
import os

# Function to generate a unique filename
def get_unique_filename(base_name, extension=".xlsx"):
    counter = 1
    file_name = f"{base_name}{extension}"
    while os.path.exists(file_name):
        file_name = f"{base_name}_{counter}{extension}"
        counter += 1
    return file_name

# Prompt for the file name, number of images, label name, and task range
file_name_base = input("Enter the base name for the file: ")
num_images = int(input("Enter the number of images: "))
label_name = input("Enter the label name for the INSTRUCTION sheet: ")

# Prompt for task start and end numbers
start_number = input("Enter the start number for the task: ")
end_number = input("Enter the end number for the task: ")

# Create task name string
task_name = f"Task Name: {start_number} to {end_number}"

# Number of images per sheet
images_per_sheet = 100

# Calculate the number of sheets needed
num_sheets = -(-num_images // images_per_sheet)  # Ceiling division

# Create a new workbook
wb = Workbook()

# Define styles
bold_font = Font(bold=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
left_align = Alignment(horizontal='left', vertical='center')
center_align = Alignment(horizontal='center', vertical='center')

# Add the INSTRUCTION sheet first
ws_instruction = wb.active
ws_instruction.title = "INSTRUCTION"

# Place the label name and task name on row 1 and 2 respectively
ws_instruction['A1'] = label_name
ws_instruction['A2'] = task_name

# Apply styles to the INSTRUCTION sheet
ws_instruction.column_dimensions['A'].width = 50
ws_instruction['A1'].font = bold_font
ws_instruction['A1'].alignment = left_align
ws_instruction['A2'].font = bold_font
ws_instruction['A2'].alignment = left_align

# Generate other sheets
for sheet_index in range(num_sheets):
    ws = wb.create_sheet(title=f"{int(start_number) + sheet_index}")
    
    # Define column headers
    headers = ["Job and ID", "Image No.", "Reviewer Name", "Remarks (Difficulties, findings and confusion)", "Status"]
    ws.append(headers)
    
    # Apply styles to header
    for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
        for cell in col:
            cell.font = bold_font
            cell.border = thin_border
            cell.alignment = center_align
    
    # Fill the worksheet
    for i in range(images_per_sheet):
        if (i % 25) == 0:
            ws.append(["JOB", i])
        elif (i % 25) == 1:
            ws.append(["ID", i])
        else:
            ws.append(["", i])
    
    # Apply bold style and alignment to all cells in the worksheet
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = bold_font
            if cell.column == 1:
                cell.alignment = left_align
            else:
                cell.alignment = center_align
    
    # Adjust column widths to fit content
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter  # Get the column letter
        for cell in column:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2  # Add some padding
        ws.column_dimensions[column_letter].width = adjusted_width

# Generate a unique filename
file_name = get_unique_filename(file_name_base)

# Save the workbook
wb.save(file_name)

print(f"Excel file '{file_name}' created successfully!")
